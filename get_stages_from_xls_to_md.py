from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
workbook = load_workbook(filename="TdF Stages Riders Results (2011-).xlsx", read_only=True)

def year_stages(y):
    ws = workbook[f'TdF20{str(y)} Stages']
    print(f'TdF20{str(y)} Stages')
    res_md = f"## {ws['A1'].value}\n"
    res_md += '| Stage | Type | Date | Start | Finish | Distance |\n'
    res_md += '| --- | --- | --- | --- | --- | --- |\n'
    for i in range(3, ws.max_row+1):
        r = ws[i]
        d = datetime.strptime(str(r[2].value), '%Y-%m-%d %H:%M:%S').date().strftime('%A %-d %B %Y')
        if isinstance(r[0].value, int):
            ii = f'Stage {r[0].value}'
        else:
            ii = f'{r[0].value}'
        if y != 22:
            # if r[3] contains ' > ' then split and insert the second part between [3] and [4]
            # if r[3] does not contain ' > ' then copy r[3] to a new entry between [3] and [4]
            if r[3].value.find(' > ') != -1:
                nr = r[3].value.split(' > ')
                md = f"| {ii} | {r[1].value} | {d} | {nr[0]} | {nr[1]} | {r[4].value} |\n"
            else:
                md = f"| {ii} | {r[1].value} | {d} | {r[3].value} | {r[3].value} | {r[4].value} |\n"        
        else:
            md = f"| {ii} | {r[1].value} | {d} | {r[3].value} | {r[4].value} | {r[5].value} |\n"
        #print(md)
        res_md += md
    return res_md

#with open('stages_file.md', 'w') as resfile:
#    for y in range(12, 24):
#        print(year_results(f'TdF20{str(y)} Stages'))
#        resfile.write(year_stages(y))
#        resfile.write('\n\n')

# the full resolved path of this folder this file is in
# print(Path(__file__).resolve().parent)

stagespath = f'{Path(__file__).resolve().parent}/stages'
Path(stagespath).mkdir(exist_ok=True) # make the log folder if necessary

for y in range(12, 24):
    with open(f'{stagespath}/TdF_stages_20{str(y)}.md', 'w') as resfile:
        resfile.write(year_stages(y))
        resfile.write('\n\n')
